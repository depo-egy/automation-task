# class contains the following methods:
'''
(1) login
(2) comment
(3) edit excel
'''
from selenium import webdriver
import time
from time import sleep
from openpyxl import load_workbook
import generator

class Comment:

    driver = ''
    global_comment = ''
    url = ''
    def __init__(self, url, mail_address, password, comments):

        self.url = url #forum url
        self.mail_address = mail_address
        self.password = password
        self.global_comment = comments
        self.logged_in = False

    def setup(self):

        print("Setting up")
        self.driver = webdriver.Chrome('D:/Programs/chromedriver')

    def login(self):
        self.driver.get('https://www.google.com/accounts/Login?hl=en&continue=http://www.google.co.jp/')
        self.driver.find_element_by_id("identifierId").send_keys(self.mail_address)
        self.driver.find_element_by_id("identifierNext").click()
        sleep(5)
        self.driver.find_element_by_name("password").send_keys(self.password)
        self.driver.find_element_by_id("passwordNext").click()
        
    def comment(self):
        sleep(8)
        self.driver.get(self.url)
        sleep(10)
        answer_btn = self.driver.find_element_by_class_name('WriteAnswerPrimaryActionItem')
        answer_btn.click()
        local = self.global_comment

        sleep(8)
        self.driver.execute_script(""" var local = arguments[0];
        document.getElementsByClassName('content')[2].innerHTML += local;
        """,local)

        #sleep(10)
        #b1 = self.driver.find_element_by_class_name('form_buttons')

        #self.driver.find_element_by_xpath("//div[@class='form_buttons']//a[@class='submit_button']").click()

        #option with content[2]
        click = self.driver.find_elements_by_xpath("//a[@class='submit_button']")
        click[2].click()

    def edit_excel(self):
        self.book = load_workbook("excel.xlsx")
        self.ws = self.book.get_sheet_by_name('Quora')
        for cell in self.ws["A"]:
            if cell.value is None:
                print(cell.row)
                break

        today_date = time.strftime("%x")
        self.ws.cell(row=cell.row, column=1).value = today_date
        self.ws.cell(row=cell.row, column=2).value = self.url
        self.ws.cell(row=cell.row, column=3).value = self.global_comment
        self.ws.cell(row=cell.row, column=4).value = 'https://501words.net/'
        self.book.save("excel.xlsx")


    def teardown(self):

        self.driver.quit()
        print("Tearing down")



    def main(self):
        self.setup()
        self.login()
        self.comment()
        self.edit_excel()

        self.teardown()

#Comment("https://www.quora.com/What-are-popular-gifts-for-Mothers-Day","testingmd9@gmail.com", "@qwertyuiop123","Hi Buddy").main()
